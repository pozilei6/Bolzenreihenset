from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

import main


KEY_LABELS = ("System", "Variante", "Rotor", "ze/ex", "Huelse", "Linie")
UNKNOWN_KEY_VALUE = ""
CONTROLLED_PRODUCTION_SYSTEM_PATTERN = re.compile(
    r"(?<!\d)K?(1000|2000|3000|4000|6000|8000|9000)(?:RS|S)?(?!\d)",
    flags=re.IGNORECASE,
)
UNKNOWN_FIELD_PENALTIES = {
    "System": 12,
    "Variante": 8,
    "Rotor": 10,
    "ze/ex": 6,
    "Huelse": 7,
    "Linie": 5,
}
WORD_FILE_NOTE_PATTERN = re.compile(r"Word-Datei:\s*([^;\n]+?\.docx?)", flags=re.IGNORECASE)
PAIR_COLUMNS = tuple(
    column
    for index in range(1, 6)
    for column in (f"Bolzen {index}", f"Gegenbolzen {index}")
)
DISPLAY_COLUMNS = (
    "Kandidat",
    "Wahrscheinlichkeit",
    "Status",
    "Code",
    *PAIR_COLUMNS,
    "Widerspruch",
    "Hinweis",
)


@dataclass(frozen=True)
class ArticleRecord:
    row_number: int
    description: str | None
    article: str
    prefix: str | None
    systems: tuple[str, ...]
    keys: tuple[tuple[str, str, str, str, str, str], ...]


@dataclass(frozen=True)
class CalculationRow:
    row_number: int
    article: str
    code: str
    pairs: tuple[tuple[str | None, str | None], ...]


@dataclass
class Candidate:
    article: ArticleRecord
    source: str
    status: str
    probability: int
    note: str
    has_exact_system_match: bool = False
    word_files: set[str] = field(default_factory=set)
    codes: dict[str, list[tuple[str | None, str | None]]] = field(default_factory=lambda: defaultdict(list))
    contradiction: str = ""
    details: list[str] = field(default_factory=list)


def normalized_article(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).casefold()


def clean_display(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def resolve_project_path(path) -> Any:
    return path if path.is_absolute() else Path(__file__).resolve().parent / path


def limited_pairs(pairs: list[tuple[str | None, str | None]]) -> tuple[tuple[str | None, str | None], ...]:
    result: list[tuple[str | None, str | None]] = []
    seen: set[tuple[str | None, str | None]] = set()

    for bolzen_article, gegenbolzen_article in pairs:
        pair = (bolzen_article, gegenbolzen_article)
        if pair in seen:
            continue
        seen.add(pair)
        result.append(pair)
        if len(result) == 5:
            break

    while len(result) < 5:
        result.append((None, None))

    return tuple(result)


def has_any_pair(pairs: tuple[tuple[str | None, str | None], ...]) -> bool:
    return any(bolzen_article or gegenbolzen_article for bolzen_article, gegenbolzen_article in pairs)


def controlled_production_system_numbers(values: tuple[str | None, ...]) -> set[str]:
    numbers: set[str] = set()
    for value in values:
        if value is None:
            continue
        numbers.update(match.group(1) for match in CONTROLLED_PRODUCTION_SYSTEM_PATTERN.finditer(str(value)))
    return numbers


def production_system_number_matches(record: ArticleRecord, word_keys: tuple[tuple[str, str, str, str, str, str], ...]) -> bool:
    record_numbers = controlled_production_system_numbers(record.systems)
    if not record_numbers:
        record_numbers = controlled_production_system_numbers(tuple(key[0] for key in record.keys))
    if not record_numbers:
        return True

    word_numbers = controlled_production_system_numbers(tuple(key[0] for key in word_keys))
    return bool(record_numbers & word_numbers)


def production_system_exactly_matches(record: ArticleRecord, word_keys: tuple[tuple[str, str, str, str, str, str], ...]) -> bool:
    record_systems = {key[0] for key in record.keys if key[0]}
    if not record_systems:
        record_systems = set(record.systems)

    return bool(record_systems & {key[0] for key in word_keys if key[0]})


def is_unknown_key_value(value: str) -> bool:
    return value == UNKNOWN_KEY_VALUE


def key_similarity(
    word_keys: tuple[tuple[str, str, str, str, str, str], ...],
    target_keys: tuple[tuple[str, str, str, str, str, str], ...],
) -> tuple[int, tuple[str, ...], tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    best_score = -1
    best_diff: tuple[str, ...] = tuple(KEY_LABELS)
    best_unknown: tuple[str, ...] = tuple(KEY_LABELS)
    best_word_key: tuple[str, ...] = tuple(UNKNOWN_KEY_VALUE for _ in KEY_LABELS)
    best_target_key: tuple[str, ...] = tuple(UNKNOWN_KEY_VALUE for _ in KEY_LABELS)

    for word_key in word_keys:
        for target_key in target_keys:
            diff = []
            unknown = []
            score = 0
            for label, word_value, target_value in zip(KEY_LABELS, word_key, target_key):
                if is_unknown_key_value(word_value):
                    unknown.append(label)
                elif word_value == target_value:
                    score += 1
                else:
                    diff.append(label)

            diff_fields = tuple(diff)
            unknown_fields = tuple(unknown)
            if (
                score > best_score
                or (score == best_score and len(diff_fields) < len(best_diff))
                or (
                    score == best_score
                    and len(diff_fields) == len(best_diff)
                    and len(unknown_fields) < len(best_unknown)
                )
            ):
                best_score = score
                best_diff = diff_fields
                best_unknown = unknown_fields
                best_word_key = word_key
                best_target_key = target_key

    return max(best_score, 0), best_diff, best_unknown, best_word_key, best_target_key


def format_key_differences(
    diff_fields: tuple[str, ...],
    word_key: tuple[str, ...],
    target_key: tuple[str, ...],
) -> str:
    if not diff_fields:
        return ""

    parts = []
    for field in diff_fields:
        index = KEY_LABELS.index(field)
        word_value = word_key[index] or "-"
        target_value = target_key[index] or "-"
        parts.append(f"{field}: Word={word_value}, Excel={target_value}")

    return "; ".join(parts)


def extract_values_after_label(pattern: str, text: str) -> str | None:
    match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return None

    value = match.group(1)
    value = re.split(r"[\x07\r\n]", value, maxsplit=1)[0]
    return main.normalize_text(value)


def raw_variant(filename: str, system: str | None, function: str | None, text: str) -> str:
    combined = " ".join(part or "" for part in (filename, system, function))

    if re.search(r"FP\s*2", combined, flags=re.IGNORECASE):
        return "FP2"
    if re.search(r"SS04|SP[_\s.-]*0[.,]4|(?<!\d)0[.,]4(?!\d)", combined, flags=re.IGNORECASE):
        return "SP_0.4"
    if re.search(r"SS05|SP[_\s.-]*0[.,]5|(?<!\d)0[.,]5(?!\d)", combined, flags=re.IGNORECASE):
        return "SP_0.5"
    if re.search(r"\bVario\b|Change\s*Code|ChangeCode", combined, flags=re.IGNORECASE):
        return "VARIO"
    if re.search(r"OM\s*3|Omega\s*3", combined, flags=re.IGNORECASE):
        return "OM3"
    if re.search(r"OM\s*2|Omega\s*2", combined, flags=re.IGNORECASE):
        return "OM2"
    if main.is_4000s_casa_word(system, filename):
        return "CASA"

    return main.word_match_variant({"Datei": filename, "System": system, "Funktion": function})


def raw_rotor_parts(value: str | None) -> tuple[str, str]:
    try:
        return main.parse_rotor_diameter_and_type(value)
    except ValueError:
        text = (main.normalize_text(value) or "").lower()
        if re.search(r"exzentr|(?<![a-z])ex(?![a-z])", text):
            return UNKNOWN_KEY_VALUE, "ex"
        if re.search(r"zentr|(?<![a-z])ze(?![a-z])", text):
            return UNKNOWN_KEY_VALUE, "ze"
        return UNKNOWN_KEY_VALUE, UNKNOWN_KEY_VALUE


def raw_hulse_diameters(value: str | None) -> tuple[str, ...]:
    try:
        return tuple(main.normalize_diameters(value, "Hulsen-diam"))
    except ValueError:
        return (UNKNOWN_KEY_VALUE,)


def raw_line_value(system: str, value: str | None, filename: str) -> str:
    try:
        return main.normalize_line(value)
    except ValueError:
        filename_line = main.line_from_filename(filename)
        if filename_line:
            return filename_line

        default_line = main.default_line_for_system(system)
        if default_line:
            return default_line[0]

        return UNKNOWN_KEY_VALUE


def raw_word_match_keys(word_data: dict[str, Any]) -> tuple[tuple[str, str, str, str, str, str], ...]:
    filename = word_data.get("Datei") or ""
    word_path = main.DOCX_FOLDER / filename
    if not word_path.exists():
        return tuple()

    text = main.read_word_text(word_path).replace("\x00", " ")
    system_value = word_data.get("System") or extract_values_after_label(r"System\s*:\s*([^\x07\r\n]+)", text)
    rotor_value = word_data.get("Rotor-diam") or extract_values_after_label(r"Rotor[^0-9]{0,20}([^\x07\r\n]+)", text)
    hulse_value = word_data.get("Hulsen-diam") or extract_values_after_label(r"H[\w\W]{0,4}lsen[^0-9]{0,20}([^\x07\r\n]+)", text)
    line_value = word_data.get("Linie") or extract_values_after_label(r"Linie\s*:\s*([^\x07\r\n]+)", text) or main.line_from_filename(filename)
    function_value = word_data.get("Funktion") or extract_values_after_label(r"Funktion\s*:\s*([^\x07\r\n]+)", text)

    try:
        system = main.normalize_system(system_value or main.system_from_filename(filename))
    except ValueError:
        return tuple()

    rotor_diameter, rotor_type = raw_rotor_parts(rotor_value)
    hulse_diameters = raw_hulse_diameters(hulse_value)
    line = raw_line_value(system, line_value, filename)
    variant = raw_variant(filename, system_value, function_value, text)
    return tuple(
        (system, variant, rotor_diameter, rotor_type, hulse_diameter, line)
        for hulse_diameter in hulse_diameters
    )


class ReadOnlyRepository:
    def __init__(self) -> None:
        self.bolzenreihenset_pairs = main.load_bolzenreihenset_pairs()
        self.article_records = self._load_article_records()
        self.article_records_by_normalized = self._index_article_records()
        self.calculation_rows_by_article, self.reverse_pair_index = self._load_calculation_rows()
        self.word_data = main.load_all_word_data()
        self.word_filename_by_reference = self._index_word_filenames_by_reference()
        self.reference_by_word_filename = self._index_references_by_word_filename()
        self.fill_tables_by_filename: dict[str, dict[str, list[dict[str, Any]]]] = {}
        self.word_key_cache: dict[str, tuple[tuple[str, str, str, str, str, str], ...]] = {}

        workbook = load_workbook(main.XLSX_FILE, read_only=True, data_only=True)
        try:
            self.drawing_lookup = main.load_drawing_lookup(workbook)
            self.reference_by_excel_row = self._load_references_by_excel_row(workbook)
        finally:
            workbook.close()

    def _load_article_records(self) -> list[ArticleRecord]:
        records: list[ArticleRecord] = []
        for row_number, pair in self.bolzenreihenset_pairs.items():
            description, article = pair
            if not article:
                continue

            article_text = str(article).strip()
            production_system = main.production_system_from_description(description)
            systems = tuple(sorted({production_system} if production_system else set()))
            records.append(
                ArticleRecord(
                    row_number=row_number,
                    description=None if description is None else str(description).strip(),
                    article=article_text,
                    prefix=main.article_prefix(article_text),
                    systems=systems,
                    keys=tuple(main.parse_excel_match_keys(pair, include_se_matches=True)),
                )
            )

        return records

    def _index_word_filenames_by_reference(self) -> dict[str, str]:
        return {
            str(word_data["Referenz_zu_Abfullnorm"]).strip(): str(word_data["Datei"]).strip()
            for word_data in self.word_data
            if word_data.get("Referenz_zu_Abfullnorm") and word_data.get("Datei")
        }

    def _index_references_by_word_filename(self) -> dict[str, str]:
        return {
            str(word_data["Datei"]).strip(): str(word_data["Referenz_zu_Abfullnorm"]).strip()
            for word_data in self.word_data
            if word_data.get("Referenz_zu_Abfullnorm") and word_data.get("Datei")
        }

    def _load_references_by_excel_row(self, workbook) -> dict[int, str]:
        sheet = workbook[main.ABFUELLNORM_SHEET_NAME]
        references: dict[int, str] = {}

        for row_number, (reference,) in enumerate(
            sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True),
            start=2,
        ):
            if reference is not None and str(reference).strip():
                references[row_number] = str(reference).strip()

        return references

    def word_file_note_for_record(self, record: ArticleRecord) -> str:
        reference = self.reference_by_excel_row.get(record.row_number)
        if not reference:
            return "Word-Datei: keine Referenz in Excel."

        filename = self.word_filename_by_reference.get(reference)
        if filename:
            return f"Word-Datei: {filename}"

        return f"Word-Datei: nicht gefunden fuer Referenz {reference}."

    def _index_article_records(self) -> dict[str, list[ArticleRecord]]:
        index: dict[str, list[ArticleRecord]] = defaultdict(list)
        for record in self.article_records:
            index[normalized_article(record.article)].append(record)
        return index

    def _load_calculation_rows(
        self,
    ) -> tuple[
        dict[str, list[CalculationRow]],
        dict[tuple[str | None, str, str, str, str, str], set[str]],
    ]:
        rows_by_article: dict[str, list[CalculationRow]] = defaultdict(list)
        reverse_index: dict[tuple[str | None, str, str, str, str, str], set[str]] = defaultdict(set)
        systems_by_article = main.production_systems_by_bolzenreihenset(self.bolzenreihenset_pairs)

        workbook = load_workbook(main.XLSX_FILE, read_only=True, data_only=False)
        try:
            sheet = workbook[main.CALC_SHEET_NAME]
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                article = row[1] if len(row) > 1 else None
                code = row[2] if len(row) > 2 else None
                if article is None or code is None:
                    continue

                article_text = str(article).strip()
                code_text = str(code).strip()
                pairs: list[tuple[str | None, str | None]] = []
                for offset in range(3, 13, 2):
                    bolzen_article = row[offset] if offset < len(row) else None
                    gegenbolzen_article = row[offset + 1] if offset + 1 < len(row) else None
                    pair = (
                        None if bolzen_article is None else str(bolzen_article).strip(),
                        None if gegenbolzen_article is None else str(gegenbolzen_article).strip(),
                    )
                    pairs.append(pair)

                    if pair[0] and pair[1]:
                        for system in systems_by_article.get(article_text, set()):
                            reverse_index[(main.article_prefix(article_text), system, code_text, pair[0], pair[1])].add(article_text)

                rows_by_article[normalized_article(article_text)].append(
                    CalculationRow(
                        row_number=row_number,
                        article=article_text,
                        code=code_text,
                        pairs=limited_pairs(pairs),
                    )
                )
        finally:
            workbook.close()

        return rows_by_article, reverse_index

    def fill_tables(self, word_filename: str) -> dict[str, list[dict[str, Any]]]:
        if word_filename not in self.fill_tables_by_filename:
            self.fill_tables_by_filename[word_filename] = main.parse_fill_tables(main.DOCX_FOLDER / word_filename)
        return self.fill_tables_by_filename[word_filename]

    def word_keys(self, word_data: dict[str, Any]) -> tuple[tuple[str, str, str, str, str, str], ...]:
        filename = word_data.get("Datei") or ""
        if filename not in self.word_key_cache:
            try:
                keys = tuple(main.parse_word_match_keys(word_data))
            except ValueError:
                keys = raw_word_match_keys(word_data)
            self.word_key_cache[filename] = keys
        return self.word_key_cache[filename]

    def article_systems(self, record: ArticleRecord) -> tuple[str, ...]:
        if record.systems:
            return record.systems

        return tuple(sorted({key[0] for key in record.keys}))


@lru_cache(maxsize=1)
def repository() -> ReadOnlyRepository:
    return ReadOnlyRepository()


def make_display_row(
    candidate: str,
    probability: int,
    status: str,
    code: str,
    pairs: tuple[tuple[str | None, str | None], ...],
    note: str,
    contradiction: str = "",
) -> dict[str, str]:
    row = {
        "Kandidat": candidate,
        "Wahrscheinlichkeit": f"{probability} %",
        "Status": status,
        "Code": code,
        "Widerspruch": contradiction,
        "Hinweis": note,
    }

    for index, (bolzen_article, gegenbolzen_article) in enumerate(limited_pairs(list(pairs)), start=1):
        row[f"Bolzen {index}"] = clean_display(bolzen_article)
        row[f"Gegenbolzen {index}"] = clean_display(gegenbolzen_article)

    return row


def word_file_to_open_for_rows(rows: list[dict[str, str]]) -> dict[str, str] | None:
    for row in rows:
        note = row.get("Hinweis", "")
        match = WORD_FILE_NOTE_PATTERN.search(note)
        if not match:
            continue

        filename = match.group(1).strip()
        path = resolve_project_path(main.DOCX_FOLDER / filename)
        if path.exists():
            return {"filename": filename, "path": str(path)}

    return None


def display_rows_for_existing(record: ArticleRecord, rows: list[CalculationRow], word_file_note: str) -> list[dict[str, str]]:
    display_rows = []
    for calculation_row in rows:
        if has_any_pair(calculation_row.pairs):
            display_rows.append(
                make_display_row(
                    record.article,
                    100,
                    "Excel vorhanden",
                    calculation_row.code,
                    calculation_row.pairs,
                    f"Berechnungscode-Zeile {calculation_row.row_number}; Bolzenpaar im Excel vorhanden. {word_file_note}",
                )
            )
            continue

        row = make_display_row(
            record.article,
            0,
            "nicht vorhanden",
            calculation_row.code,
            calculation_row.pairs,
            f"Berechnungscode-Zeile {calculation_row.row_number}; im Excel steht fuer diesen Code kein Bolzen/Gegenbolzen-Paar. {word_file_note}",
        )
        row["Wahrscheinlichkeit"] = "-"
        display_rows.append(row)

    return display_rows


def empty_existing_codes(rows: list[CalculationRow]) -> list[str]:
    return [calculation_row.code for calculation_row in rows if not has_any_pair(calculation_row.pairs)]


def probability_from_evidence(
    key_score: int,
    diff_fields: tuple[str, ...],
    unknown_fields: tuple[str, ...],
    target_votes: int,
    total_pairs: int,
    is_top_article: bool,
    top_is_tied: bool,
) -> int:
    if key_score == len(KEY_LABELS):
        probability = 96
    elif key_score == 5:
        probability = 76
    elif key_score == 4:
        probability = 62
    else:
        probability = 48

    if target_votes:
        probability += min(14, target_votes * 2)
    if total_pairs and target_votes >= max(2, total_pairs // 2):
        probability += 5
    if is_top_article:
        probability += 6
    if top_is_tied:
        probability -= 8

    for field in diff_fields:
        if field == "System":
            probability -= 12
        elif field == "Variante":
            probability -= 8
        elif field == "Huelse":
            probability -= 7
        elif field == "ze/ex":
            probability -= 6
        elif field == "Linie":
            probability -= 5
        elif field == "Rotor":
            probability -= 10

    for field in unknown_fields:
        probability -= UNKNOWN_FIELD_PENALTIES.get(field, 5)

    if unknown_fields:
        strong_reverse_anchor = target_votes >= 2 and is_top_article and not top_is_tied
        probability = min(probability, 86 if strong_reverse_anchor else 76)

    return max(35, min(99, probability))


def status_from_probability(probability: int, top_is_tied: bool, has_contradiction: bool = False) -> str:
    if has_contradiction:
        return "Widerspruch pruefen" if probability >= 60 else "unsicher"
    if probability >= 92 and not top_is_tied:
        return "sehr wahrscheinlich"
    if probability >= 78 and not top_is_tied:
        return "wahrscheinlich"
    if probability >= 60:
        return "Alternative pruefen"
    return "unsicher"


def rows_from_word_fill_tables(
    repo: ReadOnlyRepository,
    word_filename: str,
    prefix: str,
) -> tuple[dict[str, list[tuple[str | None, str | None]]], int, list[str]]:
    rows_by_code: dict[str, list[tuple[str | None, str | None]]] = defaultdict(list)
    reasons: list[str] = []
    fill_tables = repo.fill_tables(word_filename)
    table_rows = fill_tables.get(prefix, [])

    for fill_row in table_rows:
        code = str(fill_row.get("Code") or "").strip()
        if not code:
            continue

        bolzen_article, gegenbolzen_article, mapping_reasons = main.map_fill_row_to_articles(fill_row, repo.drawing_lookup)
        if mapping_reasons:
            reasons.extend(mapping_reasons)
            continue

        rows_by_code[code].append((bolzen_article, gegenbolzen_article))

    return rows_by_code, len(table_rows), reasons


def reverse_votes_for_word(
    repo: ReadOnlyRepository,
    word_data: dict[str, Any],
    record: ArticleRecord,
    prefix: str,
) -> tuple[int, int, bool, bool, Counter[str]]:
    word_filename = word_data.get("Datei") or ""
    article_systems = repo.article_systems(record)
    vote_counter: Counter[str] = Counter()
    total_pairs = 0
    target_votes = 0
    fill_tables = repo.fill_tables(word_filename)

    for fill_row in fill_tables.get(prefix, []):
        code = str(fill_row.get("Code") or "").strip()
        bolzen_article, gegenbolzen_article, reasons = main.map_fill_row_to_articles(fill_row, repo.drawing_lookup)
        if not code or not bolzen_article or not gegenbolzen_article or reasons:
            continue

        total_pairs += 1
        for system in article_systems:
            hits = repo.reverse_pair_index.get((prefix, system, code, bolzen_article, gegenbolzen_article), set())
            for article in hits:
                vote_counter[article] += 1
                if normalized_article(article) == normalized_article(record.article):
                    target_votes += 1

    if not vote_counter:
        return target_votes, total_pairs, False, False, vote_counter

    highest_vote = vote_counter.most_common(1)[0][1]
    top_articles = [article for article, count in vote_counter.items() if count == highest_vote]
    is_top_article = any(normalized_article(article) == normalized_article(record.article) for article in top_articles)
    top_is_tied = len(top_articles) > 1
    return target_votes, total_pairs, is_top_article, top_is_tied, vote_counter


def collect_word_candidates(repo: ReadOnlyRepository, record: ArticleRecord) -> list[Candidate]:
    if not record.prefix or not record.keys:
        return []

    candidates: list[Candidate] = []

    for word_data in repo.word_data:
        word_filename = word_data.get("Datei") or ""
        word_keys = repo.word_keys(word_data)
        if not word_keys:
            continue
        if not production_system_number_matches(record, word_keys):
            continue
        has_exact_system_match = production_system_exactly_matches(record, word_keys)

        key_score, diff_fields, unknown_fields, word_key, target_key = key_similarity(word_keys, record.keys)
        contradiction = format_key_differences(diff_fields, word_key, target_key)
        target_votes, total_pairs, is_top_article, top_is_tied, vote_counter = reverse_votes_for_word(
            repo,
            word_data,
            record,
            record.prefix,
        )

        include_candidate = key_score >= 5 or (key_score >= 4 and target_votes) or target_votes >= 2
        if not include_candidate:
            continue

        rows_by_code, table_row_count, mapping_reasons = rows_from_word_fill_tables(repo, word_filename, record.prefix)
        if not rows_by_code:
            continue

        probability = probability_from_evidence(
            key_score,
            diff_fields,
            unknown_fields,
            target_votes,
            total_pairs,
            is_top_article,
            top_is_tied,
        )
        status = status_from_probability(probability, top_is_tied, bool(diff_fields))

        note_parts = [
            f"Word-Datei: {word_filename}",
            f"Key-Treffer: {key_score}/6",
        ]
        if diff_fields:
            note_parts.append("abweichend: " + ", ".join(diff_fields))
        if unknown_fields:
            note_parts.append("fehlend: " + ", ".join(unknown_fields))
        if target_votes:
            note_parts.append(f"Rueckwaertsanker: {target_votes}/{total_pairs}")
        if is_top_article:
            note_parts.append("Top-Kandidat nach Artikelpaaren")
        if top_is_tied:
            note_parts.append("Gleichstand mit Alternativen")
        if mapping_reasons:
            note_parts.append("nicht gemappte Zeilen vorhanden")

        candidate = Candidate(
            article=record,
            source="Word/Rueckwaertsanker",
            status=status,
            probability=probability,
            note="; ".join(note_parts),
            has_exact_system_match=has_exact_system_match,
            word_files={word_filename},
            contradiction=contradiction,
        )
        candidate.codes = rows_by_code
        candidate.details = [f"Artikelpaar-Stimmen: {dict(vote_counter)}"] if vote_counter else []
        candidates.append(candidate)

    if any(candidate.has_exact_system_match for candidate in candidates):
        candidates = [candidate for candidate in candidates if candidate.has_exact_system_match]

    candidates.sort(key=lambda item: (-item.probability, sorted(item.word_files), item.note))
    return candidates


def display_rows_for_candidates(
    candidates: list[Candidate],
    max_candidates: int = 40,
    excluded_codes: set[str] | None = None,
) -> tuple[list[dict[str, str]], set[str]]:
    excluded_codes = excluded_codes or set()
    rows: list[dict[str, str]] = []
    skipped_codes: set[str] = set()
    seen: set[tuple[str, int, str, str, tuple[tuple[str | None, str | None], ...]]] = set()

    for candidate in candidates[:max_candidates]:
        for code, pairs in sorted(candidate.codes.items(), key=lambda item: item[0]):
            if code in excluded_codes:
                skipped_codes.add(code)
                continue

            visible_pairs = limited_pairs(pairs)
            identity = (candidate.article.article, candidate.probability, candidate.status, code, visible_pairs)
            if identity in seen:
                continue
            seen.add(identity)
            rows.append(
                make_display_row(
                    candidate.article.article,
                    candidate.probability,
                    candidate.status,
                    code,
                    visible_pairs,
                    candidate.note,
                    candidate.contradiction,
                )
            )

    return rows, skipped_codes


def candidate_word_file_detail_lines(
    record: ArticleRecord,
    candidates: list[Candidate],
    max_files: int = 10,
) -> list[str]:
    filenames = sorted(
        {
            filename
            for candidate in candidates
            for filename in candidate.word_files
            if filename
        }
    )
    if not filenames:
        return []

    visible_filenames = filenames[:max_files]
    lines = [
        f"{record.article}: Gefundene Word-Dateien aus Kandidatenlogik "
        "(nur Anzeige, nicht automatisch gewaehlt):"
    ]
    lines.extend(f"- {filename}" for filename in visible_filenames)

    remaining_count = len(filenames) - len(visible_filenames)
    if remaining_count:
        lines.append(f"... {remaining_count} weitere")

    return lines


def candidate_word_filename(candidate: Candidate) -> str | None:
    return next(iter(candidate.word_files), None)


def writeable_99_match_for_candidate(
    repo: ReadOnlyRepository,
    record: ArticleRecord,
    candidate: Candidate,
) -> tuple[ArticleRecord, Candidate, str, str] | None:
    if candidate.probability < 99 or candidate.contradiction:
        return None

    filename = candidate_word_filename(candidate)
    if not filename:
        return None

    reference = repo.reference_by_word_filename.get(filename)
    if not reference:
        return None

    return record, candidate, filename, reference


def collect_writeable_99_matches(
    repo: ReadOnlyRepository,
    records: list[ArticleRecord],
) -> tuple[list[tuple[ArticleRecord, Candidate, str, str]], list[str]]:
    record_candidates: list[tuple[ArticleRecord, Candidate, str, str]] = []
    errors: list[str] = []

    for record in records:
        for candidate in collect_word_candidates(repo, record):
            match = writeable_99_match_for_candidate(repo, record, candidate)
            if match:
                record_candidates.append(match)
                continue

            if candidate.probability >= 99 and not candidate.contradiction:
                filename = candidate_word_filename(candidate)
                if filename and not repo.reference_by_word_filename.get(filename):
                    errors.append(f"{record.article}: keine Referenz zur Word-Datei {filename} gefunden.")

    return record_candidates, errors


def has_single_writeable_99_word_file(record_candidates: list[tuple[ArticleRecord, Candidate, str, str]]) -> bool:
    return len({filename for _record, _candidate, filename, _reference in record_candidates}) == 1


def existing_reference_covers_base_reference(existing_reference: str, reference: str) -> bool:
    base_match = re.fullmatch(r"(800\.\d{3}\.\d{3})", reference, flags=re.IGNORECASE)
    if not base_match:
        return False

    return bool(
        re.fullmatch(
            rf"{re.escape(base_match.group(1))}\.[A-Za-z]",
            existing_reference,
            flags=re.IGNORECASE,
        )
    )


def write_99_percent_match(article_text: str) -> dict[str, Any]:
    repo = repository()
    normalized = normalized_article(article_text)
    records = repo.article_records_by_normalized.get(normalized, [])
    writes: list[dict[str, str]] = []
    messages: list[str] = []
    errors: list[str] = []

    if not article_text.strip():
        return {"writes": writes, "messages": ["Bitte ein Bolzenreihenset eingeben."], "errors": errors}

    if not records:
        return {
            "writes": writes,
            "messages": [f"Bolzenreihenset nicht exakt im Excel gefunden: {article_text}"],
            "errors": errors,
        }

    record_candidates, reference_errors = collect_writeable_99_matches(repo, records)
    errors.extend(reference_errors)

    if not record_candidates:
        return {
            "writes": writes,
            "messages": ["Kein widerspruchsfreier 99 %-Treffer zum Schreiben gefunden."],
            "errors": errors,
        }

    filenames = {filename for _record, _candidate, filename, _reference in record_candidates}
    if len(filenames) > 1:
        errors.append("Mehrere 99 %-Word-Dateien gefunden: " + ", ".join(sorted(filenames)))
        return {"writes": writes, "messages": messages, "errors": errors}

    workbook = load_workbook(main.XLSX_FILE)
    try:
        abfuellnorm_sheet = workbook[main.ABFUELLNORM_SHEET_NAME]
        calc_sheet = workbook[main.CALC_SHEET_NAME]
        calc_rows = main.calculation_rows_by_article_and_code(calc_sheet)
        systems_by_article = main.production_systems_by_bolzenreihenset(repo.bolzenreihenset_pairs)
        report_items: list[dict[str, Any]] = []
        pair_writes: list[dict[str, Any]] = []

        for record, candidate, filename, reference in record_candidates:
            reference_cell = abfuellnorm_sheet[f"C{record.row_number}"]
            current_reference = None if reference_cell.value is None else str(reference_cell.value).strip()
            if current_reference in {None, ""}:
                reference_cell.value = reference
                writes.append(
                    {
                        "sheet": main.ABFUELLNORM_SHEET_NAME,
                        "cell": reference_cell.coordinate,
                        "value": reference,
                    }
                )
            elif current_reference != reference and not existing_reference_covers_base_reference(current_reference, reference):
                errors.append(
                    f"{main.ABFUELLNORM_SHEET_NAME}!{reference_cell.coordinate} enthaelt {current_reference}, "
                    f"nicht {reference}."
                )

            production_system = main.production_system_from_description(record.description)
            for code, pairs in sorted(candidate.codes.items(), key=lambda item: item[0]):
                calculation_row_numbers, system_error = main.find_calculation_rows(
                    calc_rows,
                    record.article,
                    code,
                    production_system,
                    systems_by_article,
                )
                if system_error:
                    errors.append(f"{record.article} Code {code}: {system_error}")
                    continue

                if not calculation_row_numbers:
                    calculation_row_number = main.append_calculation_code_row(
                        calc_sheet,
                        calc_rows,
                        record.article,
                        code,
                    )
                    writes.extend(
                        [
                            {
                                "sheet": main.CALC_SHEET_NAME,
                                "cell": f"A{calculation_row_number}",
                                "value": calc_sheet[f"A{calculation_row_number}"].value,
                            },
                            {
                                "sheet": main.CALC_SHEET_NAME,
                                "cell": f"B{calculation_row_number}",
                                "value": record.article,
                            },
                            {
                                "sheet": main.CALC_SHEET_NAME,
                                "cell": f"C{calculation_row_number}",
                                "value": code,
                            },
                        ]
                    )
                elif len(calculation_row_numbers) != 1:
                    errors.append(
                        f"{record.article} Code {code}: Berechnungscode-Zeile nicht eindeutig "
                        f"({calculation_row_numbers})."
                    )
                    continue
                else:
                    calculation_row_number = calculation_row_numbers[0]

                context = {
                    "Referenz_zu_Abfullnorm": reference,
                    "Word_Datei": filename,
                    "Excel_Artikel": record.article,
                    "Produktionssystem": production_system,
                    "Tabelle": record.prefix,
                    "Code": code,
                    "Berechnungscode_Zeile": calculation_row_number,
                }
                for bolzen_article, gegenbolzen_article in limited_pairs(list(pairs)):
                    main.write_pair_to_next_available_slot(
                        calc_sheet,
                        calculation_row_number,
                        bolzen_article,
                        gegenbolzen_article,
                        context,
                        pair_writes,
                        report_items,
                    )

        for item in pair_writes:
            writes.extend(
                [
                    {
                        "sheet": main.CALC_SHEET_NAME,
                        "cell": item["Bolzen_Zelle"],
                        "value": item["Bolzen_Artikel"],
                    },
                    {
                        "sheet": main.CALC_SHEET_NAME,
                        "cell": item["Gegenbolzen_Zelle"],
                        "value": item["Gegenbolzen_Artikel"],
                    },
                ]
            )

        for item in report_items:
            errors.append(
                f"{item.get('Excel_Artikel')} Code {item.get('Code')}: {item.get('Grund')}"
            )

        if writes:
            workbook.save(main.XLSX_FILE)
            messages.append(f"{len(writes)} Zellen geschrieben.")
            clear_cache()
        else:
            messages.append("Nichts geschrieben; alle passenden Werte waren bereits vorhanden.")
    finally:
        workbook.close()

    return {"writes": writes, "messages": messages, "errors": errors}


def get_display_rows(article_text: str) -> dict[str, Any]:
    repo = repository()
    normalized = normalized_article(article_text)
    records = repo.article_records_by_normalized.get(normalized, [])

    if not article_text.strip():
        return {
            "columns": DISPLAY_COLUMNS,
            "rows": [],
            "messages": ["Bitte ein Bolzenreihenset eingeben."],
            "details": [],
            "has_writeable_99_matches": False,
        }

    if not records:
        return {
            "columns": DISPLAY_COLUMNS,
            "rows": [],
            "messages": [f"Bolzenreihenset nicht exakt im Excel gefunden: {article_text}"],
            "details": ["Es wird nichts geschrieben. Bitte Schreibweise in Spalte B pruefen."],
            "has_writeable_99_matches": False,
        }

    display_rows: list[dict[str, str]] = []
    messages: list[str] = []
    details: list[str] = []
    writeable_99_matches: list[tuple[ArticleRecord, Candidate, str, str]] = []

    for record in records:
        messages.append(
            f"Treffer: Excelzeile {record.row_number}, {record.description or 'ohne Beschreibung'}"
        )

        existing_rows = repo.calculation_rows_by_article.get(normalized_article(record.article), [])
        existing_codes_with_pairs = {
            calculation_row.code
            for calculation_row in existing_rows
            if has_any_pair(calculation_row.pairs)
        }
        word_file_note = repo.word_file_note_for_record(record)
        display_rows.extend(display_rows_for_existing(record, existing_rows, word_file_note))

        candidates = collect_word_candidates(repo, record)
        writeable_99_matches.extend(
            match
            for candidate in candidates
            if (match := writeable_99_match_for_candidate(repo, record, candidate))
        )
        candidate_rows, skipped_codes = display_rows_for_candidates(candidates, excluded_codes=existing_codes_with_pairs)
        display_rows.extend(candidate_rows)

        non_empty_existing_rows = [calculation_row for calculation_row in existing_rows if has_any_pair(calculation_row.pairs)]
        blank_codes = empty_existing_codes(existing_rows)
        if non_empty_existing_rows:
            details.append(f"{record.article}: {len(non_empty_existing_rows)} vorhandene Berechnungscode-Zeilen mit Bolzenpaaren gelesen.")
        if blank_codes:
            details.append(
                f"{record.article}: Berechnungscode-Zeilen ohne Bolzenpaar als 'nicht vorhanden' angezeigt: {', '.join(blank_codes)}."
            )
        if candidates:
            details.append(f"{record.article}: {len(candidates)} wahrscheinliche Word-Kandidaten gefunden.")
            details.extend(candidate_word_file_detail_lines(record, candidates))
        if skipped_codes:
            details.append(
                f"{record.article}: Word-Vorschlaege fuer bereits belegte Excel-Codes ausgeblendet: {', '.join(sorted(skipped_codes))}."
            )

    if not display_rows:
        messages.append("Keine Bolzen/Gegenbolzen-Zeilen gefunden. Bitte Excel und Word-Daten pruefen.")

    return {
        "columns": DISPLAY_COLUMNS,
        "rows": display_rows,
        "messages": messages,
        "details": details,
        "word_file_to_open": word_file_to_open_for_rows(display_rows),
        "has_writeable_99_matches": has_single_writeable_99_word_file(writeable_99_matches),
    }


def clear_cache() -> None:
    repository.cache_clear()


if __name__ == "__main__":
    import json
    import sys

    query = " ".join(sys.argv[1:]).strip()
    if not query:
        query = "HK_12/29_ze_L2_W_78_D_2.4_RHOM"

    print(json.dumps(get_display_rows(query), ensure_ascii=False, indent=2))
